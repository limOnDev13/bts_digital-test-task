import io
import json
import random
from datetime import datetime, timedelta
from string import ascii_uppercase, digits
from typing import Dict

from django.http import JsonResponse
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .factories import RobotFactory
from .models import Robot

SYMBOLS: str = ascii_uppercase + digits


def _return_random_model() -> str:
    """Returns random model (version)."""
    return random.choice(SYMBOLS) + random.choice(SYMBOLS)


class CreateRobotTestCase(TestCase):
    """Test case for create_robot"""

    test_robots_data = [
        {
            "model": "R2",
            "version": "D2",
            "created": "2022-12-31 23:59:00",
        },
        {
            "model": "A3",
            "version": "XS",
            "created": "2023-01-01 00:00:00",
        },
        {
            "model": "X5",
            "version": "LT",
            "created": "2023-01-01 00:00:01",
        },
    ]

    def setUp(self):
        self.random_robot_data: Dict[str, str] = {
            "model": random.choice(ascii_uppercase) + random.choice(digits),
            "version": _return_random_model(),
            "created": str(datetime.now()),
        }
        self.test_robots_data.append(self.random_robot_data)

    def tearDown(self):
        self.test_robots_data.remove(self.random_robot_data)

    def test_create_robot(self):
        """Test the creation of a new robot entity."""
        for test_data in self.test_robots_data:
            response: JsonResponse = self.client.post(
                reverse("robots:create_robots"),
                json.dumps(test_data),
                content_type="application/json",
            )
            self.assertEqual(response.status_code, 201)
            self.assertTrue(
                Robot.objects.filter(
                    model=test_data["model"],
                    version=test_data["version"],
                    created=test_data["created"],
                ).exists()
            )

    def test_create_robot_with_not_existing_model(self):
        """A negative test of creating a robot with a non-existent model."""
        invalid_data: Dict[str, str] = {
            "model": "11",
            "version": "11",
            "created": "2023-01-01 00:00:01",
        }
        response: JsonResponse = self.client.post(
            reverse("robots:create_robots"),
            json.dumps(invalid_data),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)


class GetSummaryTestCase(TestCase):
    """Test case for view func get_summary_in_file."""

    def setUp(self):
        self.old_robots = [
            RobotFactory.create(
                created=datetime.now() - timedelta(days=random.randint(10, 20))
            )
            for _ in range(random.randint(20, 50))
        ]
        self.new_robots = [
            RobotFactory.create(created=datetime.now())
            for _ in range(random.randint(20, 50))
        ]

    def tearDown(self):
        for robot in self.old_robots:
            robot.delete()
        for robot in self.new_robots:
            robot.delete()

    def test_get_summary_in_file(self):
        """Test getting a summary about produced robots."""
        response = self.client.get(reverse("robots:summary"))
        self.assertEqual(response.status_code, 200)

        content = response.content
        wb = load_workbook(filename=io.BytesIO(content))

        count_robots_in_excel: int = 0
        for ws in wb:
            self.assertEqual(ws["A1"].value, "Модель")
            self.assertEqual(ws["B1"].value, "Версия")
            self.assertEqual(ws["C1"].value, "Количество за неделю")
            for cell in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                self.assertEqual(cell[0], ws.title)
            count_robots_in_excel += sum(
                int(cell[0])
                for cell in ws.iter_rows(
                    min_row=2, min_col=3, max_col=3, values_only=True
                )
            )
        self.assertEqual(count_robots_in_excel, len(self.new_robots))
