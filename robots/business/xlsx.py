"""The module responsible for generating a summary in the format .xlsx."""

from typing import List, Set

from openpyxl import Workbook

from .models import ProducedRobots, ProducedRobotsSummary


class ProducedRobotsXLSX(ProducedRobotsSummary):
    """A class for saving the summary in the format .xlsx."""

    fields = ("Модель", "Версия", "Количество за неделю")

    def __init__(self, summary: List[ProducedRobots]):
        super().__init__(summary)
        self.wb = Workbook()

    def create_sheets_with_headers(self):
        """Create pages for each model and record column headings."""
        models: Set[str] = {robot.model for robot in self.produced_robots}
        for model in models:
            ws = self.wb.create_sheet(model)
            ws.append(self.fields)

        if models:
            # remove default sheet
            self.wb.remove(self.wb["Sheet"])

    def write_in_workbook(self):
        """Write data to the table."""
        for robot in self.produced_robots:
            self.wb[robot.model].append(robot.data_in_list)

    def save_file(self, tmp) -> None:
        """Save summary in file .xlsx."""
        self.create_sheets_with_headers()
        self.write_in_workbook()
        self.wb.save(tmp.name)
